import os
import sys
import requests
from bs4 import BeautifulSoup
from PIL import Image
from pyzbar.pyzbar import decode
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
import tkinter as tk
from tkinter import filedialog, messagebox
from concurrent.futures import ThreadPoolExecutor, as_completed
import time


def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def get_website_title(url, max_retries=3):
    for attempt in range(max_retries):
        try:
            response = requests.get(url, timeout=10)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')
            title = soup.title.string if soup.title else "No title found"
            return title.strip() if title else "No title found"
        except requests.RequestException as e:
            if attempt == max_retries - 1:
                return f"Unable to fetch title: {str(e)}"
            time.sleep(1)  # Wait for 1 second before retrying


def process_image(image_path):
    try:
        with Image.open(image_path) as img:
            decoded_objects = decode(img)
            results = []
            for obj in decoded_objects:
                data = obj.data.decode('utf-8')
                company_name = get_website_title(data)
                results.append({
                    'filename': os.path.basename(image_path),
                    'data': data,
                    'company': company_name
                })
            return results
    except Exception as e:
        print(f"Error processing {image_path}: {str(e)}")
        return []


def read_qr_codes(folder_path):
    qr_data = []
    image_files = [f for f in os.listdir(folder_path) if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
    total_files = len(image_files)

    with ThreadPoolExecutor(max_workers=10) as executor:
        future_to_file = {executor.submit(process_image, os.path.join(folder_path, file)): file for file in image_files}
        for i, future in enumerate(as_completed(future_to_file), 1):
            qr_data.extend(future.result())
            print(f"Processed {i}/{total_files} files")

    return qr_data


def save_to_excel(data, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "QR Code Data"

    headers = ['Filename', 'Company', 'QR Code Data', 'Apply']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for row, item in enumerate(data, start=2):
        ws.cell(row=row, column=1, value=item['filename'])
        ws.cell(row=row, column=2, value=item['company'])

        cell = ws.cell(row=row, column=3, value=item['data'])
        cell.hyperlink = item['data']
        cell.font = Font(color="0000FF", underline="single")

    dv = DataValidation(type="list", formula1='"☐,☑"', allow_blank=True)
    ws.add_data_validation(dv)

    for row in range(2, len(data) + 2):
        cell = ws.cell(row=row, column=4, value="☐")
        dv.add(cell)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    try:
        wb.save(output_file)
    except PermissionError:
        return False
    return True


def main():
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    messagebox.showinfo("QR Code Reader",
                        "Welcome to the QR Code Reader!\n\nPlease select the folder containing QR code images.")
    folder_path = filedialog.askdirectory(title="Select Folder with QR Code Images")

    if not folder_path:
        messagebox.showerror("Error", "No folder selected. Exiting.")
        return

    messagebox.showinfo("QR Code Reader", "Please select the output folder.")
    output_folder = filedialog.askdirectory(title="Select Output Folder")

    if not output_folder:
        messagebox.showerror("Error", "No output folder selected. Exiting.")
        return

    messagebox.showinfo("QR Code Reader", "Please enter the desired output filename (including .xlsx extension).")
    output_filename = filedialog.asksaveasfilename(
        initialdir=output_folder,
        title="Save Excel File",
        filetypes=[("Excel files", "*.xlsx")],
        defaultextension=".xlsx"
    )

    if not output_filename:
        messagebox.showerror("Error", "No filename selected. Exiting.")
        return

    messagebox.showinfo("QR Code Reader",
                        "Processing QR codes. This may take a while depending on the number of images and network speed.")

    qr_data = read_qr_codes(folder_path)

    save_success = save_to_excel(qr_data, output_filename)
    if not save_success:
        messagebox.showerror("Error",
                             f"Unable to save the file. Please ensure that {output_filename} is not open in another program and you have write permissions.")
        return

    messagebox.showinfo("QR Code Reader", f"QR code data has been extracted and saved to:\n{output_filename}")


if __name__ == "__main__":
    main()