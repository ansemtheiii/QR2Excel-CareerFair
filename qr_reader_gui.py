import os
import re
from PIL import Image
from pyzbar.pyzbar import decode
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
import tkinter as tk
from tkinter import filedialog


def extract_company_name(url):
    match = re.search(r'https?://(?:www\.)?([^/]+)', url)
    if match:
        domain = match.group(1)
        company = re.split(r'[.-]', domain.rsplit('.', 2)[0])[0]
        return company.capitalize()
    return "Unknown"


def read_qr_codes(folder_path):
    qr_data = []
    for filename in os.listdir(folder_path):
        if filename.endswith(('.png', '.jpg', '.jpeg')):
            image_path = os.path.join(folder_path, filename)
            try:
                with Image.open(image_path) as img:
                    decoded_objects = decode(img)
                    for obj in decoded_objects:
                        data = obj.data.decode('utf-8')
                        company_name = extract_company_name(data)
                        qr_data.append({
                            'filename': filename,
                            'data': data,
                            'company': company_name
                        })
            except Exception as e:
                print(f"Error processing {filename}: {str(e)}")
    return qr_data


def save_to_excel(data, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "QR Code Data"

    headers = ['Filename', 'Company', 'QR Code Data', 'Apply']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for row, item in enumerate(data, start=2):
        ws.cell(row=row, column=1, value=item['filename'])
        ws.cell(row=row, column=2, value=item['company'])

        cell = ws.cell(row=row, column=3, value=item['data'])
        cell.hyperlink = item['data']
        cell.font = Font(color="0000FF", underline="single")

    dv = DataValidation(type="list", formula1='"☐,☑"', allow_blank=True)
    ws.add_data_validation(dv)

    for row in range(2, len(data) + 2):
        cell = ws.cell(row=row, column=4, value="☐")
        dv.add(cell)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_file)


def main():
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    print("Please select the folder containing QR code images.")
    folder_path = filedialog.askdirectory(title="Select Folder with QR Code Images")

    if not folder_path:
        print("No folder selected. Exiting.")
        return

    print("Please select the output folder.")
    output_folder = filedialog.askdirectory(title="Select Output Folder")

    if not output_folder:
        print("No output folder selected. Exiting.")
        return

    print("Please enter the desired output filename (including .xlsx extension).")
    output_filename = filedialog.asksaveasfilename(
        initialdir=output_folder,
        title="Save Excel File",
        filetypes=[("Excel files", "*.xlsx")],
        defaultextension=".xlsx"
    )

    if not output_filename:
        print("No filename selected. Exiting.")
        return

    qr_data = read_qr_codes(folder_path)
    save_to_excel(qr_data, output_filename)

    print(f"QR code data has been extracted and saved to {output_filename}")


if __name__ == "__main__":
    main()