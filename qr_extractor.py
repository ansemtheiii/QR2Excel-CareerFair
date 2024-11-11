import os
import re
from PIL import Image
from pyzbar.pyzbar import decode
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation


def extract_company_name(url):
    # Simple regex to extract domain name, which might be the company name
    match = re.search(r'https?://(?:www\.)?([^/]+)', url)
    if match:
        domain = match.group(1)
        # Remove common TLDs and split by dots or dashes
        company = re.split(r'[.-]', domain.rsplit('.', 2)[0])[0]
        return company.capitalize()
    return "Unknown"


def read_qr_codes(folder_path):
    qr_data = []
    for filename in os.listdir(folder_path):
        if filename.endswith(('.png', '.jpg', '.jpeg')):
            image_path = os.path.join(folder_path, filename)
            try:
                with Image.open(image_path) as img:
                    decoded_objects = decode(img)
                    for obj in decoded_objects:
                        data = obj.data.decode('utf-8')
                        company_name = extract_company_name(data)
                        qr_data.append({
                            'filename': filename,
                            'data': data,
                            'company': company_name
                        })
            except Exception as e:
                print(f"Error processing {filename}: {str(e)}")
    return qr_data


def save_to_excel(data, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "QR Code Data"

    # Set up headers
    headers = ['Filename', 'Company', 'QR Code Data', 'Apply']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # Add data
    for row, item in enumerate(data, start=2):
        ws.cell(row=row, column=1, value=item['filename'])
        ws.cell(row=row, column=2, value=item['company'])

        # Add hyperlink
        cell = ws.cell(row=row, column=3, value=item['data'])
        cell.hyperlink = item['data']
        cell.font = Font(color="0000FF", underline="single")

    # Add checkbox for 'Apply' column
    dv = DataValidation(type="list", formula1='"☐,☑"', allow_blank=True)
    ws.add_data_validation(dv)

    for row in range(2, len(data) + 2):
        cell = ws.cell(row=row, column=4, value="☐")
        dv.add(cell)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_file)


def main():
    folder_path = r"C:\Users\ansem\OneDrive\GATECH\Job Fair\qr_images"
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_file = os.path.join(script_dir, "qr_data.xlsx")

    qr_data = read_qr_codes(folder_path)
    save_to_excel(qr_data, output_file)

    print(f"QR code data has been extracted and saved to {output_file}")


if __name__ == "__main__":
    main()